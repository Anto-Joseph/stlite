import streamlit as st
import pandas as pd
import sqlite3
from openpyxl import load_workbook
import time

file_df = []
Xl_Sheet = 0
Xl_selectedSheetDF = 0

# --- DATABASE CONNECTION ---

con = sqlite3.connect("database-1.db")
cursor =  con.cursor()

#----- TITLE -----
st.title("Employee Master Data")

# ----- Table Data ---- 
st.session_state["EmpTbl"] = pd.read_sql_query(f"SELECT * from EMP", con)
st.session_state.EmpTbl['EmpStatus'] = st.session_state.EmpTbl['EmpStatus'].map({1: True, 0: False})

edited_df = st.data_editor(
    st.session_state["EmpTbl"],
    key = "EmpTbl_key",
    use_container_width = True,
    #on_change =  process_changes,
    num_rows="dynamic",
    column_config= {
        "EmpID" : st.column_config.TextColumn("Employee Id", 
                                                      #width = 'Small', 
                                                      required=True,
                                                    ),
        "EmpName" : st.column_config.Column("Employee Name", 
                                                      width = 'Medium', 
                                                      required=True,
                                                    ),
        "EmpDesignation" : st.column_config.SelectboxColumn("Job Title",
                                                     width = 'Medium', 
                                                     #options= []
                                                     required=True,
                                                   ),

         "EmpStatus" : st.column_config.CheckboxColumn("Employee Status",
                                                        help="Is Employye Working On Site",
                                                        default = True,),
        "EmpType" : st.column_config.TextColumn("Type", default= st.session_state.get("name")),
        "EmpRole" : st.column_config.Column("Role", width = 'Medium',),

    },

    column_order=("EmpID", "EmpName", "EmpDesignation","EmpStatus","EmpType","EmpRole"),

)


# ----- BATCH APPEND EMPLOYEE DETAILS -------



uploaded_file = st.file_uploader("Batch Append Employee Details" )
if uploaded_file is not None:
    # To read file as bytes:
    bytes_data = uploaded_file.getvalue()

    try :
        file_df = load_workbook(uploaded_file)

    except Exception as e :
        st.warning("Opps wrong file format")
        st.write(e)

if file_df != []:
    Xl_Sheet = st.selectbox(
                                "Select your sheet",
                                file_df.sheetnames,
                                index=None,
                                placeholder="Select Sheet...",
                            )
    Xl_selectedSheetDF = pd.DataFrame(file_df[Xl_Sheet].values)
    Xl_selectedSheetDF.columns = Xl_selectedSheetDF.iloc[0]
    Xl_selectedSheetDF =  Xl_selectedSheetDF.iloc[(1):] 

if st.button("Upload", type="primary"):
                try:
                    Xl_selectedSheetDF.to_sql(name='EMP', if_exists="append", con = con, index=False)
                    st.success(f"Timesheet Uploaded Scucessfully By : *{st.session_state.get('name')}*", icon = '🤘')
                    st.toast('File Updated', icon='🎉')
                    time.sleep(.5)
                except sqlite3.Error as error:
                    st.warning(error, icon = '🚩')
                    st.toast('Error Uploading!{error}', icon='🚩')
                    time.sleep(.5)

