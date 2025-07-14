import streamlit as st
import datetime
import streamlit_authenticator as stauth
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import sqlite3
import uuid
import time


Xl_SkipColumns = []
Xl_FormatDataType = 0
Xl_EmpID = 0
Xl_EmpHrs = 0
T_date = 0

# --- DATABASE CONNECTION ---

con = sqlite3.connect("database-1.db")
cursor =  con.cursor()






st.title('Daily Hours')
st.write(f'Entered By : *{st.session_state.get("name")}*')
file_df = []

T_date = st.date_input("Select Input date")


# ----- Table Data ---- 
st.session_state["DailyHrsTbl"] = pd.read_sql_query(f"SELECT * from View_Daily_Timesheet WHERE DATE = '{T_date}'", con)
st.session_state.DailyHrsTbl['EmpStatus'] = st.session_state.DailyHrsTbl['EmpStatus'].map({1: True, 0: False})

# ---- DAILY TIMESHEET TABLE ------

daily_df = st.data_editor(
    st.session_state["DailyHrsTbl"],
    #key = st.session_state['EmpTaskTbl_Key'],
    use_container_width = True,
    #on_change =  process_changes,
    num_rows="dynamic",
    column_config= {
        "EmpID" : st.column_config.Column("Emp ID", width = 'Medium',),
        "EmpName" : st.column_config.Column("Employee Name", 
                                                      width = 'Medium', 
                                                      required=True,
                                                    ),

        "EnteredBy" : st.column_config.Column("Added By", ),
        "EmpDesignation" : st.column_config.Column("Job Title", ),
        "EmpStatus" : st.column_config.Column("Employee Status", ),
        "CostCentre" : st.column_config.Column("Employee C.Center", ),
        "OT_HRS" : st.column_config.Column("O.T Hrs.", ),
        "BASIC_HRS" : st.column_config.Column("Basic Hrs.", ),
        "TOTAL_HRS" : st.column_config.Column("Total Hrs.", ),


    },

    column_order=("EmpID", "EmpName","EmpDesignation","CostCentre","EmpStatus", "OT_HRS","BASIC_HRS", "TOTAL_HRS", "EnteredBy"),

)











#------ APPEND DAILY TIMESHEETS ---------
uploaded_file = st.file_uploader("Choose a file" )
if uploaded_file is not None:
    # To read file as bytes:
    bytes_data = uploaded_file.getvalue()

    try :
        file_df = load_workbook(uploaded_file)

    except Exception as e :
        st.warning("Opps wrong file format")
        st.write(e)

if file_df != []:
    Xl_Sheet = st.selectbox(
                                "Select your sheet",
                                file_df.sheetnames,
                                index=None,
                                placeholder="Select Sheet...",
                            )
    Xl_FormatDataType = st.selectbox(
                                "Load Data As",
                                ("By Sheet","By Table"),
                                index=None,
                                placeholder="Format as...",
                            )
    if Xl_FormatDataType == "By Sheet" :
        Xl_selectedSheetDF = pd.DataFrame(file_df[Xl_Sheet].values)
        Xl_SkipRowNo =  st.number_input("Header Row Number",step = 1,min_value  = 0)
        headerRow = Xl_selectedSheetDF.loc[Xl_SkipRowNo].to_frame()
        headerRow['ColNum'] = headerRow.index
        headerRow = headerRow.dropna()
        headerRow.T
        Xl_selectedSheetDF = Xl_selectedSheetDF.iloc[0:][headerRow['ColNum']]
        Xl_selectedSheetDF.columns = Xl_selectedSheetDF.iloc[Xl_SkipRowNo]
        Xl_selectedSheetDF =  Xl_selectedSheetDF.iloc[(Xl_SkipRowNo+1):] 
        Xl_selectedSheetDF

        Xl_EmpID = st.selectbox(
                                "Employee ID Coloum",
                                headerRow[Xl_SkipRowNo],
                                index=None,
                                placeholder="Select ID Coloumn...",
                            )
        Xl_EmpHrs = st.selectbox(
                                "Total Hours",
                                headerRow[Xl_SkipRowNo],
                                index=None,
                                placeholder="Select Hours Coloumn...",
        
                            )
        try : 
            Xl_selectedSheetDF = Xl_selectedSheetDF[Xl_selectedSheetDF[Xl_EmpID].notna()]
            Xl_selectedSheetDF
            DB_Data = {'EmpID' : Xl_selectedSheetDF[Xl_EmpID], 'BASIC_HRS' :Xl_selectedSheetDF[Xl_EmpHrs], 'DATE' : T_date}
            TimeSheet_forDBUplaod  =  pd.DataFrame(data=DB_Data)
            TimeSheet_forDBUplaod

            if st.button("Upload", type="primary"):
                try:
                    TimeSheet_forDBUplaod.to_sql(name='DAILYHRS', if_exists="append", con = con, index=False)
                    st.success(f"Timesheet Uploaded Scucessfully for : {T_date}, By : *{st.session_state.get('name')}*", icon = '🤘')
                    st.toast('File Updated', icon='🎉')
                    time.sleep(.5)
                except sqlite3.Error as error:
                    st.warning(error, icon = '🚩')
                    st.toast('Error Uploading!{error}', icon='🚩')
                    time.sleep(.5)
                
        except sqlite3.Error as error:
            st.warning(error, icon = '🚩')


        












